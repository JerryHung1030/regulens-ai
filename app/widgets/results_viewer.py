from __future__ import annotations

import csv # For CSV export
# import shutil # shutil.copyfile was for old report export, not needed for CSV from data.
from pathlib import Path
from typing import Optional # For type hinting
import re # Keep re for now, might be useful for _get_display_name if that's kept/adapted
import json # Added for placeholder theme colors
import openpyxl # For Excel export
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side # For Excel styling

from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QStyle, # Already present, but QToolButton needs it too
    QFileDialog,
    QMessageBox,
    QTableWidget,
    QToolButton, # Added for collapsible sections
    QTableWidgetItem,
    QHeaderView,
    QDialog,
    QPlainTextEdit,
    QSizePolicy,
    QFrame, # Added for StatsBarChart
    QScrollArea # Added for RunEvidenceDetailsDialog
)
from PySide6.QtCore import Signal, Qt, QSize # Added QSize
from PySide6.QtGui import QTextOption, QColor, QFontMetrics # Added QColor and QFontMetrics

from app.models.project import CompareProject
from app.models.docs import ExternalRegulationClause, AuditTask # For type hinting in new dialog
from app.pipeline.pipeline_v1_1 import ProjectRunData, _load_run_json # For loading results
from app.logger import logger
from app.utils.font_manager import get_font, get_display_font


# Helper function for eliding text
def elide_text(text: str | None, max_length: int = 50) -> str:
    if text is None:
        return "N/A"
    if len(text) > max_length:
        return f"{text[:max_length - 3]}..."
    return text

# StatsBarChart Class Definition START

class StatsBarChart(QWidget):
    def __init__(self, translator, parent: QWidget | None = None):
        super().__init__(parent)
        self.translator = translator
        self.setObjectName("statsBarChart")
        self.setFixedHeight(40) # Adjust height as needed

        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(2) # Spacing between segments

        self.compliant_frame = QFrame(self)
        self.compliant_frame.setProperty("status", "compliant")
        self.compliant_label = QLabel("0", self.compliant_frame)
        self.compliant_label.setAlignment(Qt.AlignCenter)
        self.compliant_label.setFont(get_display_font(size=9, weight_style='semi_bold'))

        compliant_layout = QHBoxLayout(self.compliant_frame)
        compliant_layout.addWidget(self.compliant_label)
        compliant_layout.setContentsMargins(5,0,5,0)
        self.compliant_frame.style().unpolish(self.compliant_frame) # Ensure style is reapplied
        self.compliant_frame.style().polish(self.compliant_frame)

        self.non_compliant_frame = QFrame(self)
        self.non_compliant_frame.setProperty("status", "non-compliant")

        self.non_compliant_label = QLabel("0", self.non_compliant_frame)
        self.non_compliant_label.setAlignment(Qt.AlignCenter)
        self.non_compliant_label.setFont(get_display_font(size=9, weight_style='semi_bold'))
        non_compliant_layout = QHBoxLayout(self.non_compliant_frame)
        non_compliant_layout.addWidget(self.non_compliant_label)
        non_compliant_layout.setContentsMargins(5,0,5,0)
        self.non_compliant_frame.style().unpolish(self.non_compliant_frame)
        self.non_compliant_frame.style().polish(self.non_compliant_frame)

        self.pending_frame = QFrame(self)
        self.pending_frame.setProperty("status", "pending")
        self.pending_label = QLabel("0", self.pending_frame)
        self.pending_label.setFont(get_display_font(size=9, weight_style='semi_bold'))
        self.pending_label.setAlignment(Qt.AlignCenter)
        pending_layout = QHBoxLayout(self.pending_frame)
        pending_layout.setContentsMargins(0,0,0,0) # Ensure label is within frame
        pending_layout.addWidget(self.pending_label)
        self.pending_frame.style().unpolish(self.pending_frame)
        self.pending_frame.style().polish(self.pending_frame)

        self.na_frame = QFrame(self) # New N/A frame
        self.na_frame.setProperty("status", "na")
        # self.na_frame.setStyleSheet("background-color: #6c757d;") # Placeholder QSS will handle
        self.na_label = QLabel("0", self.na_frame)
        self.na_label.setFont(get_display_font(size=9, weight_style='semi_bold'))
        self.na_label.setAlignment(Qt.AlignCenter)
        # self.na_label.setStyleSheet("color: white; font-weight: bold;") # Placeholder QSS will handle
        na_layout = QHBoxLayout(self.na_frame)
        na_layout.setContentsMargins(0,0,0,0) # Ensure label is within frame
        na_layout.addWidget(self.na_label)
        self.na_frame.style().unpolish(self.na_frame)
        self.na_frame.style().polish(self.na_frame)
        
        self.layout.addWidget(self.compliant_frame, 0)
        self.layout.addWidget(self.non_compliant_frame, 0)
        self.layout.addWidget(self.pending_frame, 0)
        self.layout.addWidget(self.na_frame, 0) # Add N/A frame to layout

        self._update_tooltips() # Initial tooltip setup
        self.translator.language_changed.connect(self._update_tooltips)


    def setData(self, compliant_count: int, non_compliant_count: int, pending_count: int, na_count: int, total_relevant_count: int):
        self.compliant_label.setText(str(compliant_count))
        self.non_compliant_label.setText(str(non_compliant_count))
        self.pending_label.setText(str(pending_count))
        self.na_label.setText(str(na_count)) # Set N/A count

        if total_relevant_count == 0:
            self.layout.setStretchFactor(self.compliant_frame, 1) # Show one segment as placeholder
            self.layout.setStretchFactor(self.non_compliant_frame, 0)
            self.layout.setStretchFactor(self.pending_frame, 0)
            self.layout.setStretchFactor(self.na_frame, 0)
            
            self.compliant_frame.setVisible(True) 
            self.compliant_label.setText(self.translator.get("stats_bar_no_data", "No data"))
            self.non_compliant_frame.setVisible(False)
            self.pending_frame.setVisible(False)
            self.na_frame.setVisible(False)
        else:
            self.compliant_frame.setVisible(compliant_count > 0)
            self.non_compliant_frame.setVisible(non_compliant_count > 0)
            self.pending_frame.setVisible(pending_count > 0)
            self.na_frame.setVisible(na_count > 0) # Set visibility for N/A

            self.layout.setStretchFactor(self.compliant_frame, compliant_count if compliant_count > 0 else 0)
            self.layout.setStretchFactor(self.non_compliant_frame, non_compliant_count if non_compliant_count > 0 else 0)
            self.layout.setStretchFactor(self.pending_frame, pending_count if pending_count > 0 else 0)
            self.layout.setStretchFactor(self.na_frame, na_count if na_count > 0 else 0) # Set stretch for N/A
        
        self._update_tooltips(compliant_count, non_compliant_count, pending_count, na_count)
        self.layout.activate()

    def _update_tooltips(self, compliant: int = 0, non_compliant: int = 0, pending: int = 0, na: int = 0): # Added na parameter
        self.compliant_frame.setToolTip(self.translator.get("stats_tooltip_compliant", "Compliant: {count}").format(count=compliant))
        self.non_compliant_frame.setToolTip(self.translator.get("stats_tooltip_non_compliant", "Non-Compliant: {count}").format(count=non_compliant))
        self.pending_frame.setToolTip(self.translator.get("stats_tooltip_pending", "Pending: {count}").format(count=pending))
        self.na_frame.setToolTip(self.translator.get("stats_tooltip_na", "N/A: {count}").format(count=na)) # Tooltip for N/A

# StatsBarChart Class Definition END

class RunEvidenceDetailsDialog(QDialog):
    """
    New dialog to display details from ProjectRunData:
    ExternalRegulationClause text, ALL its AuditTasks, their top_k evidence, and judge reasoning.
    """
    def __init__(self, clause: ExternalRegulationClause, translator, parent: QWidget | None = None): # task parameter removed
        super().__init__(parent)
        self.clause = clause
        self.translator = translator # Store translator

        # 外規標題優先順序: title > text > id
        self.clause_title_display = clause.title if getattr(clause, 'title', None) else (clause.text if getattr(clause, 'text', None) else clause.id)
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.setFixedSize(750, 650) # Increased dialog size for more content

        main_dialog_layout = QVBoxLayout(self) # This is the dialog's main layout

        # Scroll Area Setup
        scroll_area = QScrollArea(self)
        scroll_area.setWidgetResizable(True)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)

        content_widget = QWidget() # Widget to hold all scrollable content
        content_layout = QVBoxLayout(content_widget) # Layout for the content_widget

        # ExternalRegulation Clause Text
        self.clause_label = QLabel() # Text set in _retranslate_ui
        self.clause_label.setFont(get_display_font(size=11, weight_style='semi_bold'))
        content_layout.addWidget(self.clause_label)
        self.clause_text_edit = QPlainTextEdit(clause.text)
        self.clause_text_edit.setFont(get_display_font(size=10))
        self.clause_text_edit.setReadOnly(True)
        self.clause_text_edit.setFixedHeight(100) # Keep fixed height, it has its own scroll
        content_layout.addWidget(self.clause_text_edit)

        # --- Section 2: Analysis Results (Overall Clause Assessment) ---
        self.analysis_results_title_label = QLabel() # Key: "dialog_analysis_results_title"
        self.analysis_results_title_label.setFont(get_display_font(size=12, weight_style='bold'))
        content_layout.addWidget(self.analysis_results_title_label)

        self.overall_status_label = QLabel() # Key: "dialog_overall_clause_status_label" + status
        self.overall_status_label.setFont(get_display_font(size=10))
        self.overall_status_label.setWordWrap(True)
        content_layout.addWidget(self.overall_status_label)
        
        self.overall_description_label = QLabel() # Key: "dialog_overall_clause_description_label"
        self.overall_description_label.setFont(get_display_font(size=10, weight_style='semi_bold')) # Make it a sub-heading
        content_layout.addWidget(self.overall_description_label)
        self.overall_description_text = QPlainTextEdit() # Actual description
        self.overall_description_text.setFont(get_display_font(size=10))
        self.overall_description_text.setReadOnly(True)
        self.overall_description_text.setFixedHeight(80) # Adjust as needed
        content_layout.addWidget(self.overall_description_text)

        self.overall_suggestions_label = QLabel() # Key: "dialog_overall_clause_suggestions_label"
        self.overall_suggestions_label.setFont(get_display_font(size=10, weight_style='semi_bold')) # Make it a sub-heading
        content_layout.addWidget(self.overall_suggestions_label)
        self.overall_suggestions_text = QPlainTextEdit() # Actual suggestions
        self.overall_suggestions_text.setFont(get_display_font(size=10))
        self.overall_suggestions_text.setReadOnly(True)
        self.overall_suggestions_text.setFixedHeight(80) # Adjust as needed
        content_layout.addWidget(self.overall_suggestions_text)

        # --- Section 3: Evidence Details (Per Audit Task) ---
        self.evidence_section_title_label = QLabel() # Key: "dialog_evidence_section_title"
        self.evidence_section_title_label.setFont(get_display_font(size=12, weight_style='bold'))
        content_layout.addWidget(self.evidence_section_title_label)
        
        # Container for all tasks (this will be the "Evidence Section" content)
        self.tasks_container_widget = QWidget()
        self.tasks_container_layout = QVBoxLayout(self.tasks_container_widget)
        self.tasks_container_layout.setContentsMargins(0, 0, 0, 0) # No extra margins for the container itself
        self.tasks_container_layout.setSpacing(10) 
        content_layout.addWidget(self.tasks_container_widget)
        
        self.all_tasks_widgets_list = [] 

        self.no_tasks_label = QLabel() # Placeholder if no tasks for evidence section
        self.no_tasks_label.setObjectName("noTasksLabelForDialog")
        self.no_tasks_label.setTextFormat(Qt.RichText)
        self.no_tasks_label.setFont(get_display_font(size=10))
        self.no_tasks_label.setVisible(False) 
        # This will be added to tasks_container_layout in _retranslate_ui if needed
        
        content_layout.addStretch(1) 
        content_widget.setLayout(content_layout)
        scroll_area.setWidget(content_widget)
        main_dialog_layout.addWidget(scroll_area, 1)

        self.ok_button = QPushButton()
        self.ok_button.setObjectName("okButton")
        self.ok_button.setFont(get_display_font(size=10))
        self.ok_button.clicked.connect(self.accept)

        btn_h_layout = QHBoxLayout()
        btn_h_layout.addStretch()
        btn_h_layout.addWidget(self.ok_button)
        main_dialog_layout.addLayout(btn_h_layout) # Add button layout to main_dialog_layout
        
        self.translator.language_changed.connect(self._retranslate_ui)
        self._retranslate_ui() # Initial translation

    def _toggle_evidence_item(self, checked: bool, button: QToolButton, details_widget: QWidget):
        details_widget.setVisible(checked)
        # button.setArrowType(Qt.DownArrow if checked else Qt.RightArrow) # Replaced by setIcon
        if checked:
            button.setIcon(self.style().standardIcon(QStyle.SP_ArrowDown))
        else:
            button.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
        # Attempting to color the icon green via stylesheet is tricky for standard icons.
        # The stylesheet on the button itself would be:
        # button.setStyleSheet("QToolButton { icon: green; }") /* This might not work for standard icons */
        # Or more specifically targeting the icon sub-control if available.
        # For now, we'll rely on the default icon color, or a theme might provide a green variant.
        # A more robust way for specific green icons would be to load green PNGs.

    def _retranslate_ui(self):
        # 1. Clear previous dynamic task and evidence items
        for task_widget_dict in self.all_tasks_widgets_list:
            if task_widget_dict.get('task_toggle_button'): task_widget_dict['task_toggle_button'].deleteLater()
            if task_widget_dict.get('task_details_container'): task_widget_dict['task_details_container'].deleteLater()
            # Evidence items within task_details_container will be deleted with their parent.
        self.all_tasks_widgets_list.clear()
        
        # Also clear any direct children of self.tasks_container_layout not in all_tasks_widgets_list (like no_tasks_label if added directly)
        while self.tasks_container_layout.count():
            item = self.tasks_container_layout.takeAt(0)
            widget = item.widget()
            if widget:
                widget.deleteLater()

        # 2. Set Window Title and Clause Info (Static part)
        self.setWindowTitle(self.translator.get("run_evidence_details_title_clause", "Details for Clause: {clause_title}").format(clause_title=self.clause_title_display))
        self.clause_label.setText(f"<b>{self.translator.get('external_regulation_clause_heading', 'ExternalRegulation Clause:')} {self.clause.id} - {self.clause_title_display}</b>")

        # 3. Populate Section 2: Analysis Results (Overall Clause Assessment)
        self.analysis_results_title_label.setText(f"<b>{self.translator.get('dialog_analysis_results_title', 'Analysis Results (Overall Clause Assessment)')}</b>")
        
        clause_compliant_status_val = self.clause.metadata.get('clause_compliant')
        overall_status_text_display = self.translator.get('n_a', "N/A") # Default
        if clause_compliant_status_val is True:
            overall_status_text_display = self.translator.get("compliant_true_status", "Compliant")
        elif clause_compliant_status_val is False:
            overall_status_text_display = self.translator.get("compliant_false_status", "Non-Compliant")
        elif self.clause.need_procedure is False : # If no procedure needed, can be considered N/A if not explicitly judged
             overall_status_text_display = self.translator.get("compliant_na_status", "N/A (Procedure Not Required)")
        elif clause_compliant_status_val is None: # Pending or not judged but procedure was needed
            overall_status_text_display = self.translator.get("compliant_pending_status", "Pending")

        self.overall_status_label.setText(f"<b>{self.translator.get('dialog_overall_clause_status_label', 'Overall Clause Compliance Status:')}</b> {overall_status_text_display}")
        
        self.overall_description_label.setText(f"<b>{self.translator.get('dialog_overall_clause_description_label', 'Overall Clause Compliance Description:')}</b>")
        self.overall_description_text.setPlainText(self.clause.metadata.get('clause_compliance_description', self.translator.get('reasoning_not_available', 'N/A')))
        
        self.overall_suggestions_label.setText(f"<b>{self.translator.get('dialog_overall_clause_suggestions_label', 'Overall Clause Improvement Suggestions:')}</b>")
        self.overall_suggestions_text.setPlainText(self.clause.metadata.get('clause_improvement_suggestions', self.translator.get('reasoning_not_available', 'N/A')))

        # 4. Populate Section 3: Evidence Details (Per Audit Task)
        self.evidence_section_title_label.setText(f"<b>{self.translator.get('dialog_evidence_section_title', 'Evidence Details (Per Audit Task)')}</b>")

        if not self.clause.tasks:
            self.no_tasks_label.setText(f"<i>{self.translator.get('no_audit_tasks_for_clause', 'No audit tasks defined for this clause.')}</i>")
            self.tasks_container_layout.addWidget(self.no_tasks_label)
            self.no_tasks_label.setVisible(True)
        else:
            self.no_tasks_label.setVisible(False) 
            for task_idx, task_data in enumerate(self.clause.tasks):
                task_widgets = {}

                # --- Level 1 Toggle: AuditTask Sentence ---
                task_toggle_button = QToolButton()
                task_toggle_button.setText(f"{self.translator.get('audit_task_label', 'Audit Task')} {task_data.id}: {task_data.sentence}")
                task_toggle_button.setCheckable(True)
                task_toggle_button.setChecked(False) 
                task_toggle_button.setFont(get_display_font(size=10, weight_style='semi_bold')) # Normal size for task sentence
                task_toggle_button.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
                # task_toggle_button.setArrowType(Qt.RightArrow) # Replaced by setIcon
                task_toggle_button.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
                task_toggle_button.setProperty("isToggle", True) # For QSS styling
                task_toggle_button.setStyleSheet("""
                    QToolButton[isToggle=true] { border: none; text-align: left; padding: 4px; }
                    QToolButton[isToggle=true]::indicator { image: none; } /* Attempt to remove native arrow if any */
                    /* Standard icon coloring is hard with QSS. This is a best effort.
                       A more reliable method is using custom (e.g. SVG) icons. */
                    QToolButton[isToggle=true] { icon-size: 12px; /* Adjust size if needed */ } 
                    /* Placeholder for green icon - this specific QSS might not work directly on standard icons */
                    /* QToolButton[isToggle=true] { qproperty-icon: url(path/to/green/arrow.svg); } */
                    /* Forcing color on standard icons is difficult. We primarily rely on the setIcon call. */
                """)

                self.tasks_container_layout.addWidget(task_toggle_button)
                task_widgets['task_toggle_button'] = task_toggle_button
                
                task_evidence_container = QWidget() # This container is for evidence items of this task
                task_evidence_container.setVisible(False) 
                task_evidence_layout = QVBoxLayout(task_evidence_container)
                task_evidence_layout.setContentsMargins(20, 5, 0, 5) 
                task_evidence_layout.setSpacing(5)
                self.tasks_container_layout.addWidget(task_evidence_container)
                task_widgets['task_details_container'] = task_evidence_container # Re-using key for consistency in cleanup

                task_toggle_button.toggled.connect(
                    lambda checked, btn=task_toggle_button, det=task_evidence_container: self._toggle_evidence_item(checked, btn, det)
                )

                # REMOVED: Task-specific compliance status, description, suggestions.
                # These are now shown at the clause level in Section 2.

                # --- Evidence List for this Task (within task_evidence_container) ---
                # evidence_heading_for_task_label was here, can be removed or kept simple
                # For simplicity, we'll let the evidence toggles speak for themselves.
                # If needed, a small label "Found X evidence items:" can be added here.

                if task_data.top_k:
                    for ev_idx, ev_item in enumerate(task_data.top_k):
                        # --- Level 2 Toggle: Evidence Item Summary ---
                        score_val = ev_item.get('score', 0.0)
                        score_str = f"{score_val:.4f}" if isinstance(score_val, float) else str(score_val)
                        source_display_name = elide_text(ev_item.get('source_txt', self.translator.get('n_a', 'N/A')), 30) 
                        page_no = ev_item.get('page_no', self.translator.get('n_a', 'N/A'))

                        evidence_toggle_text = self.translator.get('evidence_item_title_template_dialog', "Evidence {item_num}: {source_file} (Page: {page}, Score: {score})").format(
                            item_num=ev_idx + 1, source_file=source_display_name, page=page_no, score=score_str
                        )
                        evidence_toggle_button = QToolButton()
                        evidence_toggle_button.setText(evidence_toggle_text)
                        evidence_toggle_button.setCheckable(True)
                        evidence_toggle_button.setChecked(False) 
                        evidence_toggle_button.setFont(get_display_font(size=10, weight_style='normal'))
                        evidence_toggle_button.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
                        # evidence_toggle_button.setArrowType(Qt.RightArrow) # Replaced by setIcon
                        evidence_toggle_button.setIcon(self.style().standardIcon(QStyle.SP_ArrowRight))
                        evidence_toggle_button.setProperty("isToggle", True) # For QSS styling
                        evidence_toggle_button.setStyleSheet("""
                            QToolButton[isToggle=true] { border: none; text-align: left; padding: 3px; margin-left: 0px; }
                            QToolButton[isToggle=true]::indicator { image: none; }
                            QToolButton[isToggle=true] { icon-size: 12px; }
                        """)
                        task_evidence_layout.addWidget(evidence_toggle_button)
                        
                        # --- Level 3 Content: Evidence Details (Excerpt) ---
                        full_source_file = ev_item.get('source_txt', self.translator.get('n_a', 'N/A')) 
                        excerpt = ev_item.get('excerpt', self.translator.get('n_a', 'N/A'))
                        evidence_details_html = (
                             f"<b>{self.translator.get('source_label', 'Source')}:</b> {full_source_file}<br>"
                             f"<b>{self.translator.get('page_label', 'Page')}:</b> {page_no}<br>"
                             f"<b>{self.translator.get('score_label', 'Score')}:</b> {score_str}<br>"
                             f"<b>{self.translator.get('excerpt_label', 'Excerpt')}:</b><br><i>{excerpt}</i>"
                        )
                        evidence_detail_label = QLabel(evidence_details_html)
                        evidence_detail_label.setFont(get_display_font(size=10))
                        evidence_detail_label.setWordWrap(True)
                        evidence_detail_label.setTextFormat(Qt.RichText)
                        evidence_detail_label.setVisible(False) 
                        evidence_detail_label.setStyleSheet("QLabel { padding-left: 20px; background-color: transparent; }") # Indent L3 content
                        task_evidence_layout.addWidget(evidence_detail_label)

                        evidence_toggle_button.toggled.connect(
                            lambda checked, btn=evidence_toggle_button, det=evidence_detail_label: self._toggle_evidence_item(checked, btn, det)
                        )
                else:
                    no_evidence_for_task_label = QLabel(f"<i>{self.translator.get('no_evidence_found_for_task_message', 'No evidence found for this specific task.')}</i>")
                    no_evidence_for_task_label.setFont(get_display_font(size=10))
                    # no_evidence_for_task_label.setStyleSheet("QLabel { margin-left: 10px; }") # No extra indent if it's the only thing
                    task_evidence_layout.addWidget(no_evidence_for_task_label)
                
                self.all_tasks_widgets_list.append(task_widgets)
        
        self.ok_button.setText(self.translator.get("ok_button_text", "OK"))
        logger.debug("RunEvidenceDetailsDialog UI retranslated with distinct Analysis and Evidence sections.")


class EvidenceDetailsDialog(QDialog): # Old dialog
    def __init__(self, evidence_data: dict, translator, parent: QWidget | None = None): # Added translator
        super().__init__(parent)
        self.evidence_data = evidence_data
        self.translator = translator # Store translator
        # Title set in _retranslate_ui
        self.setAttribute(Qt.WA_DeleteOnClose)

        layout = QVBoxLayout(self)

        self.lbl_file_title = QLabel() # Text set in _retranslate_ui
        self.lbl_file_title.setFont(get_display_font(size=11, weight_style='semi_bold'))
        layout.addWidget(self.lbl_file_title)

        self.lbl_analysis_title = QLabel() # Text set in _retranslate_ui
        self.lbl_analysis_title.setFont(get_display_font(size=10, weight_style='semi_bold'))
        layout.addWidget(self.lbl_analysis_title)
        
        self.txt_analysis = QPlainTextEdit()
        self.txt_analysis.setPlainText(self.evidence_data.get('analysis', 'N/A'))
        self.txt_analysis.setFont(get_display_font(size=10))
        self.txt_analysis.setReadOnly(True)
        self.txt_analysis.setWordWrapMode(QTextOption.WrapAtWordBoundaryOrAnywhere)
        layout.addWidget(self.txt_analysis)

        self.lbl_suggestion_title = QLabel() # Text set in _retranslate_ui
        self.lbl_suggestion_title.setFont(get_display_font(size=10, weight_style='semi_bold'))
        layout.addWidget(self.lbl_suggestion_title)
        
        self.txt_suggestion = QPlainTextEdit()
        self.txt_suggestion.setPlainText(self.evidence_data.get('suggestion', 'N/A'))
        self.txt_suggestion.setFont(get_display_font(size=10))
        self.txt_suggestion.setReadOnly(True)
        # txt_suggestion_text was a typo, should be self.txt_suggestion
        self.txt_suggestion.setWordWrapMode(QTextOption.WrapAtWordBoundaryOrAnywhere)
        layout.addWidget(self.txt_suggestion)

        self.btn_ok = QPushButton() # Text set in _retranslate_ui
        self.btn_ok.setObjectName("okButton")
        self.btn_ok.setFont(get_display_font(size=10))
        self.btn_ok.clicked.connect(self.accept)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)
        btn_layout.addWidget(self.btn_ok)
        layout.addLayout(btn_layout)

        self.resize(500, 400)
        
        self.translator.language_changed.connect(self._retranslate_ui)
        self._retranslate_ui() # Initial translation

    def _retranslate_ui(self):
        self.setWindowTitle(self.translator.get("evidence_details_title", "Evidence Details"))
        self.lbl_file_title.setText(f"<b>{self.translator.get('file', 'File')}:</b> {self.evidence_data.get('evidence_display_name', 'N/A')}")
        self.lbl_analysis_title.setText(self.translator.get("analysis_label", "Analysis:"))
        self.lbl_suggestion_title.setText(self.translator.get("suggestion_label", "Suggestion:"))
        self.btn_ok.setText(self.translator.get("ok_button_text", "OK")) # Reusing ok_button_text
        logger.debug("EvidenceDetailsDialog UI retranslated")


class ResultsViewer(QWidget):
    edit_requested = Signal(CompareProject)

    def __init__(self, project: CompareProject, translator, parent: QWidget | None = None): # Added translator
        super().__init__(parent)
        logger.info(f"Initializing ResultsViewer for project: {project.name}")
        self.project = project
        self.translator = translator # Store translator
        self._build_ui()
        self.project.changed.connect(self._refresh) # _refresh will call _retranslate_ui
        
        self.translator.language_changed.connect(self._retranslate_ui)
        # _retranslate_ui() # Called by _refresh via _build_ui
        logger.debug("ResultsViewer initialization completed")

    def _show_evidence_details_dialog(self, clause_id: str, _task_id: Optional[str] = None): # task_id is no longer used here
        logger.debug(f"Showing details for ExternalRegulation Clause ID: {clause_id}")
        
        if not hasattr(self.project, 'project_run_data') or not self.project.project_run_data:
            QMessageBox.warning(self, 
                                self.translator.get("data_error_title", "Data Error"), 
                                self.translator.get("project_run_data_not_loaded_text", "Project run data not loaded."))
            return

        target_clause: Optional[ExternalRegulationClause] = None
        for c_idx, c_val in enumerate(self.project.project_run_data.external_regulation_clauses):
            if c_val.id == clause_id:
                target_clause = c_val
                break
        
        if not target_clause:
            QMessageBox.warning(self, 
                                self.translator.get("data_error_title", "Data Error"), 
                                self.translator.get("external_regulation_clause_not_found_text", "ExternalRegulation Clause {clause_id} not found.").format(clause_id=clause_id))
            return

        # The dialog will now handle displaying all tasks for the clause.
        # The specific task_id is not needed to select a single task anymore.
        dialog = RunEvidenceDetailsDialog(clause=target_clause, translator=self.translator, parent=self)
        dialog.exec_()


    def _build_ui(self):
        logger.debug("Building ResultsViewer UI (Table Layout)")
        layout = self.layout()
        if layout:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget: widget.deleteLater()
        
        lay = QVBoxLayout(self)
        lay.setContentsMargins(15, 15, 15, 15)
        lay.setSpacing(10)

        title_row = QHBoxLayout()
        self._title = QLabel() # Text set in _retranslate_ui/_refresh
        self._title.setFont(get_display_font(size=14, weight_style='bold')) # Main title font
        title_row.addWidget(self._title)
        title_row.addStretch(1)

        self.btn_export_excel = QPushButton() # New button for Excel export
        self.btn_export_excel.setObjectName("btnExportExcel")
        self.btn_export_excel.setFont(get_display_font(size=10))
        self.btn_export_excel.clicked.connect(self._export_result_excel) # Connect to new method
        title_row.addWidget(self.btn_export_excel)

        self.btn_back = QPushButton() # Text/Icon set in _retranslate_ui
        self.btn_back.setObjectName("btnBack")
        self.btn_back.setFont(get_display_font(size=10))
        self.btn_back.clicked.connect(self._go_back)
        title_row.addWidget(self.btn_back)
        lay.addLayout(title_row)

        # Statistics Summary Section
        summary_section_layout = QVBoxLayout() # Main container for stats rows
        
        # Row 1: Total ExternalRegulations and Requires Procedure (existing labels) - TO BE REMOVED
        # top_summary_row_layout = QHBoxLayout()
        # self.summary_total_external_regulations_label = QLabel() 
        # self.summary_requires_procedure_label = QLabel()
        # top_summary_row_layout.addWidget(self.summary_total_external_regulations_label)
        # top_summary_row_layout.addSpacing(20)
        # top_summary_row_layout.addWidget(self.summary_requires_procedure_label)
        # top_summary_row_layout.addStretch(1)
        # summary_section_layout.addLayout(top_summary_row_layout) # REMOVED

        # Row 2: Bar Chart for Compliance Statuses (Now effectively Row 1 of summary)
        self.stats_summary_title_label = QLabel() # Title for Stats Bar Chart
        self.stats_summary_title_label.setFont(get_display_font(size=12, weight_style='bold'))
        summary_section_layout.addWidget(self.stats_summary_title_label)

        # self.stats_bar_chart = StatsBarChart(self.translator, self) # Comment out or delete
        # summary_section_layout.addWidget(self.stats_bar_chart) # Comment out or delete

        stats_text_layout = QHBoxLayout()
        stats_text_layout.setSpacing(20) # Adjust spacing as needed

        self.summary_compliant_label = QLabel()
        self.summary_compliant_label.setFont(get_display_font(size=10)) # Use appropriate font
        stats_text_layout.addWidget(self.summary_compliant_label)

        self.summary_non_compliant_label = QLabel()
        self.summary_non_compliant_label.setFont(get_display_font(size=10))
        stats_text_layout.addWidget(self.summary_non_compliant_label)

        self.summary_pending_label = QLabel()
        self.summary_pending_label.setFont(get_display_font(size=10))
        stats_text_layout.addWidget(self.summary_pending_label)

        self.summary_na_label = QLabel()
        self.summary_na_label.setFont(get_display_font(size=10))
        stats_text_layout.addWidget(self.summary_na_label)

        stats_text_layout.addStretch(1)
        summary_section_layout.addLayout(stats_text_layout)
        
        lay.addLayout(summary_section_layout)
        lay.addSpacing(10) # Existing spacing before the table

        self.detailed_results_title_label = QLabel() # Title for Detailed Results Table
        self.detailed_results_title_label.setFont(get_display_font(size=12, weight_style='bold'))
        lay.addWidget(self.detailed_results_title_label)
        # lay.addSpacing(5) # Optional small spacing after title

        self.table_widget = QTableWidget()
        self.table_widget.setFont(get_display_font(size=10)) # Set default font for table content and headers
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectRows)
        self.table_widget.setAlternatingRowColors(True)
        self.table_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Column headers set in _retranslate_ui - "Audit Task" column removed
        self.column_headers_keys = [
            "col_external_regulation_id", "col_external_regulation_title", "col_requires_procedure",
            "col_clause_compliance_status", "col_details" # Changed col_compliance_status to col_clause_compliance_status
        ]
        self.column_headers_defaults = [
            self.translator.get("col_external_regulation_id", "ExternalRegulation ID"), 
            self.translator.get("col_external_regulation_title", "ExternalRegulation Title"), 
            self.translator.get("col_requires_procedure", "Requires Procedure?"),
            self.translator.get("col_clause_compliance_status", "Clause Compliance Status"), 
            self.translator.get("col_details", "Details")
        ]
        self.table_widget.setColumnCount(len(self.column_headers_keys))
        # self.table_widget.setHorizontalHeaderLabels(...) # Done in _retranslate_ui

        # item_font = get_font(size=10) # Changed to get_display_font and set on table_widget directly
        item_font = get_display_font(size=10) # Use this for metrics if needed
        fm = QFontMetrics(item_font)
        default_row_height = fm.height() + 10 
        self.table_widget.verticalHeader().setDefaultSectionSize(default_row_height)
        self.table_widget.verticalHeader().setVisible(False)

        lay.addWidget(self.table_widget)
        self.setLayout(lay)
        
        logger.debug("ResultsViewer UI building completed. Calling initial _refresh which includes _retranslate_ui.")
        self._refresh() # This will also call _retranslate_ui

    def _retranslate_ui(self):
        self._title.setText(f"<h2>{self.translator.get('analysis_results_title', 'Analysis Results')}</h2>")
        self.btn_export_excel.setText(self.translator.get("export_excel_button", "Export Result Excel..."))
        self.btn_export_excel.setIcon(self.style().standardIcon(QStyle.SP_DialogSaveButton))
        
        self.btn_back.setText(self.translator.get("back_to_edit_button", "Back to Edit"))
        self.btn_back.setIcon(self.style().standardIcon(QStyle.SP_ArrowBack))

        # Retranslate column headers
        translated_headers = [self.translator.get(key, default) for key, default in zip(self.column_headers_keys, self.column_headers_defaults)]
        self.table_widget.setHorizontalHeaderLabels(translated_headers)
        
        # Refresh dynamic parts of the UI that depend on language
        self._refresh_summary_labels() # Separated logic for summary
        self._refresh_table_content_text() # For texts inside table cells that need retranslation

        # Retranslate new titles
        if hasattr(self, 'stats_summary_title_label'):
            self.stats_summary_title_label.setText(self.translator.get("results_viewer_stats_title", "Compliance Summary"))
        if hasattr(self, 'detailed_results_title_label'):
            self.detailed_results_title_label.setText(self.translator.get("results_viewer_table_title", "Detailed Results"))
            
        logger.debug("ResultsViewer UI retranslated")

    def _refresh_summary_labels(self):
        # 以 external_regulation 條文為單位統計合規狀態
        total_external_regulations = 0
        compliant_count = 0
        non_compliant_count = 0
        pending_count = 0
        na_count = 0

        if self.project and hasattr(self.project, 'project_run_data') and self.project.project_run_data:
            for clause in self.project.project_run_data.external_regulation_clauses:
                total_external_regulations += 1
                # 條文合規狀態判斷
                if not clause.tasks or len(clause.tasks) == 0:
                    na_count += 1
                else:
                    task_statuses = [task.compliant for task in clause.tasks]
                    if all(t is True for t in task_statuses):
                        compliant_count += 1
                    elif any(t is False for t in task_statuses):
                        non_compliant_count += 1
                    elif all(t is None for t in task_statuses):
                        pending_count += 1
                    else:
                        na_count += 1
            project_name = self.project.name
            self._title.setText(f"<h2>{self.translator.get('analysis_results_for_project_title', 'Analysis Results for: {project_name}').format(project_name=project_name)}</h2>")
        else:
            self._title.setText(f"<h2>{self.translator.get('analysis_results_title', 'Analysis Results')}</h2>")

        self.summary_compliant_label.setText(f"<b>{self.translator.get('summary_compliant_status', 'Compliant:')}</b> {compliant_count}")
        self.summary_non_compliant_label.setText(f"<b>{self.translator.get('summary_non_compliant_status', 'Non-Compliant:')}</b> {non_compliant_count}")
        self.summary_pending_label.setText(f"<b>{self.translator.get('summary_pending_status', 'Pending:')}</b> {pending_count}")
        self.summary_na_label.setText(f"<b>{self.translator.get('summary_na_status', 'N/A:')}</b> {na_count}")
        
    def _refresh_table_content_text(self):
        # This method is responsible for re-translating texts within table cells
        # if they were set with translatable strings (e.g., "Yes", "No", "N/A", status badges)
        # It assumes the table structure (rows, columns, buttons) is already there from _refresh.
        table_font = get_font(size=10)
        for row in range(self.table_widget.rowCount()):
            # Requires Procedure? (Col 2)
            item_requires_proc = self.table_widget.item(row, 2)
            if item_requires_proc: # Check if item exists
                # Re-determine original boolean value to re-translate. This is tricky.
                # Assume a data attribute was stored or re-fetch from self.project.project_run_data
                # For simplicity, this example might not perfectly re-translate existing cell content without original data.
                # A better approach: _refresh should always rebuild cells with translated text.
                # This method is more for cases where _refresh is too heavy and only text needs update.
                # Given current _refresh, this might be redundant if _refresh is called on language change.
                # However, if _refresh is NOT called, then we need to update.
                # Let's assume _refresh IS called, so this method as-is might not be strictly needed
                # unless there are other dynamic texts not covered.
                # The "View Details" button text is set in _refresh.
                # "Yes", "No", "N/A", "Compliant", "Non-Compliant", "Pending" are set in _refresh.
                pass # Most cell content is rebuilt by _refresh

        # Update "View Details" button text if it was missed by _refresh or needs specific re-translation
        for row in range(self.table_widget.rowCount()):
            button = self.table_widget.cellWidget(row, 5) # Column 5 for Details button
            if isinstance(button, QPushButton):
                button.setText(self.translator.get("view_details_button", "View Details"))


    def _handle_table_double_click(self, model_index):
        logger.debug(f"Table double-clicked at row {model_index.row()}, but action is disabled.")
        pass


    def _go_back(self):
        logger.info("Back button clicked, emitting edit_requested signal")
        self.edit_requested.emit(self.project)

    def _refresh(self):
        logger.info("Refreshing ResultsViewer (Table Display)")
        self._retranslate_ui() # Ensure all static texts are up-to-date with current language

        # Placeholder for fetching theme colors - This will be refined later
        # In a real scenario, this would involve calling theme_manager or accessing shared theme data
        def get_placeholder_theme_colors():
            # This is a simplified way to get some colors for the subtask.
            # It doesn't represent the final mechanism for theme color access.
            # We'll just pick one theme's values (e.g., dark) as stand-ins.
            # The actual theme manager will provide the correct one at runtime.
            return {
                "status_compliant_color": "#28a745",
                "text_color_on_primary": "#ffffff",
                "status_non_compliant_color": "#dc3545",
                "text_color_on_danger": "#ffffff",
                "status_pending_color": "#ffc107",
                "text_color_on_warning": "#212529", # Dark text on yellow
                "status_na_color": "#6c757d",
                "text_color_on_disabled": "#ffffff",
                "requires_procedure_yes_color": "#28a745",
                "requires_procedure_no_color": "#6c757d"
            }
        theme_colors = get_placeholder_theme_colors()
        # End of placeholder

        self.table_widget.setRowCount(0)

        if not self.project or not hasattr(self.project, 'project_run_data') or not self.project.project_run_data:
            logger.info("No project_run_data to display yet.")
            # _refresh_summary_labels handles the case of no data
            
            self.table_widget.setRowCount(1)
            item = QTableWidgetItem(self.translator.get("no_data_available", "No analysis data available. Please run the pipeline."))
            item.setTextAlignment(Qt.AlignCenter)
            table_font = get_display_font(size=10) # Ensure this uses display_font
            item.setFont(table_font)
            self.table_widget.setItem(0, 0, item)
            self.table_widget.setSpan(0, 0, self.table_widget.rowCount(), self.table_widget.columnCount()) # Span all columns
            return

        table_font = get_display_font(size=10) # Ensure this uses display_font for all items
        current_row_index = 0
        for clause_idx, clause in enumerate(self.project.project_run_data.external_regulation_clauses):
            self.table_widget.insertRow(current_row_index)
            
            # Column 0: ExternalRegulation ID
            item_clause_id = QTableWidgetItem(clause.id)
            item_clause_id.setFont(table_font)
            self.table_widget.setItem(current_row_index, 0, item_clause_id)

            # Column 1: ExternalRegulation Title
            external_regulation_title_text = clause.text
            item_external_regulation_title = QTableWidgetItem(elide_text(external_regulation_title_text, max_length=100)) # Increased max_length
            item_external_regulation_title.setFont(table_font)
            self.table_widget.setItem(current_row_index, 1, item_external_regulation_title)

            # Column 2: Requires Procedure?
            item_requires_proc = QTableWidgetItem()
            item_requires_proc.setTextAlignment(Qt.AlignCenter)
            item_requires_proc.setFont(table_font)
            if clause.need_procedure is True:
                item_requires_proc.setText(self.translator.get("yes", "Yes"))
                item_requires_proc.setBackground(QColor(theme_colors.get("requires_procedure_yes_color", "#28a745")))
                item_requires_proc.setForeground(QColor(theme_colors.get("text_color_on_primary", "#ffffff")))
            elif clause.need_procedure is False:
                item_requires_proc.setText(self.translator.get("no", "No"))
                item_requires_proc.setBackground(QColor(theme_colors.get("requires_procedure_no_color", "#6c757d")))
                item_requires_proc.setForeground(QColor(theme_colors.get("text_color_on_primary", "#ffffff")))
            else: # N/A or None
                item_requires_proc.setText(self.translator.get("n_a", "N/A"))
                item_requires_proc.setBackground(QColor(theme_colors.get("status_na_color", "#6c757d")))
                item_requires_proc.setForeground(QColor(theme_colors.get("text_color_on_disabled", "#ffffff")))
            self.table_widget.setItem(current_row_index, 2, item_requires_proc)
            
            # Column 3: Clause Compliance Status
            item_clause_compliance_status = QTableWidgetItem()
            item_clause_compliance_status.setTextAlignment(Qt.AlignCenter)
            item_clause_compliance_status.setFont(table_font)
            
            clause_compliant = clause.metadata.get('clause_compliant') # From Step 4: Judge
            if clause_compliant is True:
                item_clause_compliance_status.setText(self.translator.get("compliant_true_long", "Compliant"))
                item_clause_compliance_status.setBackground(QColor(theme_colors.get("status_compliant_color", "#d4edda")))
                item_clause_compliance_status.setForeground(QColor(theme_colors.get("text_color_on_primary", "#155724")))
            elif clause_compliant is False:
                item_clause_compliance_status.setText(self.translator.get("compliant_false_long", "Non-Compliant"))
                item_clause_compliance_status.setBackground(QColor(theme_colors.get("status_non_compliant_color", "#f8d7da")))
                item_clause_compliance_status.setForeground(QColor(theme_colors.get("text_color_on_danger", "#721c24")))
            elif clause.need_procedure is False : # Not applicable if no procedure needed and no tasks.
                item_clause_compliance_status.setText(self.translator.get("n_a", "N/A"))
                item_clause_compliance_status.setBackground(QColor(theme_colors.get("status_na_color", "#f0f0f0")))
                item_clause_compliance_status.setForeground(QColor(theme_colors.get("text_color_on_disabled", "#000000")))
            else: # Pending or not yet judged
                item_clause_compliance_status.setText(self.translator.get("compliant_pending_long", "Pending"))
                item_clause_compliance_status.setBackground(QColor(theme_colors.get("status_pending_color", "#fff3cd")))
                item_clause_compliance_status.setForeground(QColor(theme_colors.get("text_color_on_warning", "#856404")))
            self.table_widget.setItem(current_row_index, 3, item_clause_compliance_status)

            # Column 4: Details Button
            details_button = QPushButton(self.translator.get("view_details_button", "View Details"))
            details_button.setObjectName("viewDetailsButton")
            details_button.setFont(get_display_font(size=10))
            # Pass clause.id and task_id=None. The dialog will handle showing all tasks for the clause.
            details_button.clicked.connect(
                lambda checked=False, c_id=clause.id: self._show_evidence_details_dialog(c_id, None)
            )
            self.table_widget.setCellWidget(current_row_index, 4, details_button) # Index 4 for details
            
            current_row_index += 1

        self.table_widget.resizeColumnsToContents()
        # Adjust stretching: ExternalRegulation Title (column 1) is a good candidate for stretching.
        if self.table_widget.columnCount() > 1: 
            self.table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        # No longer have audit task sentence column to stretch.
        
        logger.info(f"ResultsViewer table populated with {current_row_index} rows (one per clause). Statistics updated via _refresh_summary_labels.")
        self.table_widget.viewport().update()


    def _export_result_excel(self):
        if not self.project or not hasattr(self.project, 'project_run_data') or not self.project.project_run_data:
            QMessageBox.warning(self,
                                self.translator.get("no_data_to_export_title", "No Data to Export"),
                                self.translator.get("no_data_to_export_message", "Project run data is not loaded or available. Cannot export Excel."))
            return

        project_name_safe = re.sub(r'[^\w\s-]', '', self.project.name)
        project_name_safe = re.sub(r'[-\s]+', '_', project_name_safe).strip('-_')
        if not project_name_safe.strip():
            project_name_safe = self.translator.get("untitled_project_excel_default", "Untitled_Project")

        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename_key = "audit_results_excel_default_filename"
        default_filename_pattern = "{project_name}_audit_results_{timestamp}.xlsx"
        
        suggested_filename = self.translator.get(default_filename_key, default_filename_pattern).format(
            project_name=project_name_safe, timestamp=timestamp
        )

        target_file_path_str, selected_filter = QFileDialog.getSaveFileName(
            self,
            self.translator.get("save_excel_dialog_title", "Save Result Excel As..."),
            str(Path.home() / suggested_filename),
            self.translator.get("save_excel_dialog_filter", "Excel Files (*.xlsx);;All Files (*.*)")
        )

        if not target_file_path_str:
            return

        target_file_path = Path(target_file_path_str)
        target_file_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.translator.get("excel_sheet_title_audit_results", "Audit Results")

        # Define Styles
        header_font = Font(bold=True, name='Arial', size=12)
        cell_font = Font(name='Arial', size=11)
        compliant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Light Green
        non_compliant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Light Red
        pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Light Yellow
        na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light Grey
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = [
            self.translator.get("excel_header_clause_id", "Clause ID"),
            self.translator.get("excel_header_clause_title", "Clause Title/Text"),
            self.translator.get("excel_header_requires_procedure", "Requires Procedure?"),
            self.translator.get("excel_header_clause_compliance_status", "Overall Clause Compliance"),
            self.translator.get("excel_header_clause_compliance_desc", "Clause Compliance Description"),
            self.translator.get("excel_header_clause_improvement_sugg", "Clause Improvement Suggestions"),
            self.translator.get("excel_header_task_id", "Task ID"),
            self.translator.get("excel_header_task_sentence", "Task Sentence"),
            self.translator.get("excel_header_evidence_source", "Evidence Source"),
            self.translator.get("excel_header_evidence_page", "Page"),
            self.translator.get("excel_header_evidence_score", "Score"),
            self.translator.get("excel_header_evidence_excerpt", "Evidence Excerpt")
        ]

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 40


        current_row = 2
        for clause in self.project.project_run_data.external_regulation_clauses:
            clause_start_row = current_row
            
            clause_title_display = clause.title if getattr(clause, 'title', None) else clause.text
            
            # Clause level data
            clause_data_row = {
                headers[0]: clause.id,
                headers[1]: clause_title_display,
                headers[2]: self.translator.get("yes", "Yes") if clause.need_procedure else (self.translator.get("no", "No") if clause.need_procedure is False else self.translator.get("n_a", "N/A")),
            }

            # Overall Clause Compliance Status determination
            clause_compliant_status_val = clause.metadata.get('clause_compliant')
            clause_status_display = self.translator.get('n_a', "N/A")
            clause_status_fill = na_fill
            if clause_compliant_status_val is True:
                clause_status_display = self.translator.get("compliant_true_status", "Compliant")
                clause_status_fill = compliant_fill
            elif clause_compliant_status_val is False:
                clause_status_display = self.translator.get("compliant_false_status", "Non-Compliant")
                clause_status_fill = non_compliant_fill
            elif clause.need_procedure is False:
                 clause_status_display = self.translator.get("compliant_na_status", "N/A (Procedure Not Required)")
                 clause_status_fill = na_fill
            elif clause_compliant_status_val is None:
                clause_status_display = self.translator.get("compliant_pending_status", "Pending")
                clause_status_fill = pending_fill

            clause_data_row[headers[3]] = clause_status_display
            clause_data_row[headers[4]] = clause.metadata.get('clause_compliance_description', '')
            clause_data_row[headers[5]] = clause.metadata.get('clause_improvement_suggestions', '')

            if not clause.tasks:
                for col_num, header_name in enumerate(headers, 1):
                    cell_value = clause_data_row.get(header_name, "")
                    cell = ws.cell(row=current_row, column=col_num, value=str(cell_value))
                    cell.font = cell_font
                    cell.alignment = Alignment(vertical="top", wrap_text=True, indent=0)
                    cell.border = thin_border
                    if header_name == headers[3]: # Clause Compliance Status column
                        cell.fill = clause_status_fill
                current_row += 1
            else:
                for task_idx, task in enumerate(clause.tasks):
                    task_start_row = current_row
                    task_data_row = {
                        headers[6]: task.id,
                        headers[7]: task.sentence,
                    }
                    if not task.top_k:
                        # Write clause data then task data
                        loop_row_data = {**clause_data_row, **task_data_row}
                        for col_num, header_name in enumerate(headers, 1):
                            cell_value = loop_row_data.get(header_name, "")
                            indent_level = 1 if header_name in [headers[6], headers[7]] else 0
                            cell = ws.cell(row=current_row, column=col_num, value=str(cell_value))
                            cell.font = cell_font
                            cell.alignment = Alignment(vertical="top", wrap_text=True, indent=indent_level)
                            cell.border = thin_border
                            if header_name == headers[3]: cell.fill = clause_status_fill
                        current_row += 1
                    else:
                        for evidence_idx, ev_item in enumerate(task.top_k):
                            evidence_data_row = {
                                headers[8]: ev_item.get('source_txt', ''),
                                headers[9]: str(ev_item.get('page_no', '')),
                                headers[10]: f"{ev_item.get('score', 0.0):.4f}" if isinstance(ev_item.get('score'), float) else str(ev_item.get('score', '')),
                                headers[11]: ev_item.get('excerpt', '')
                            }
                            # Write clause data, then task data, then evidence data
                            loop_row_data = {**clause_data_row, **task_data_row, **evidence_data_row}
                            for col_num, header_name in enumerate(headers, 1):
                                cell_value = loop_row_data.get(header_name, "")
                                indent_level = 0
                                if header_name in [headers[6], headers[7]]: # Task level
                                    indent_level = 1
                                elif header_name in [headers[8], headers[9], headers[10], headers[11]]: # Evidence level
                                    indent_level = 2
                                cell = ws.cell(row=current_row, column=col_num, value=str(cell_value))
                                cell.font = cell_font
                                cell.alignment = Alignment(vertical="top", wrap_text=True, indent=indent_level)
                                cell.border = thin_border
                                if header_name == headers[3]: cell.fill = clause_status_fill # Clause status
                            current_row += 1
                    # Merge task-level cells if task spanned multiple evidence rows
                    if task_start_row < current_row -1 and task.top_k: # Only if there were evidence items
                         for task_col_idx in range(6, 8): # Columns G, H (Task ID, Task Sentence)
                            ws.merge_cells(start_row=task_start_row, start_column=task_col_idx + 1, end_row=current_row - 1, end_column=task_col_idx + 1)
                            # Reapply alignment after merge for some versions of openpyxl
                            merged_cell = ws.cell(row=task_start_row, column=task_col_idx + 1)
                            merged_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)


            # Merge clause-level cells if clause spanned multiple task/evidence rows
            if clause_start_row < current_row -1 : # Only if there were tasks or evidence rows
                for clause_col_idx in range(6): # Columns A to F (Clause ID to Clause Impr. Sugg.)
                    ws.merge_cells(start_row=clause_start_row, start_column=clause_col_idx + 1, end_row=current_row - 1, end_column=clause_col_idx + 1)
                    # Reapply alignment after merge
                    merged_cell = ws.cell(row=clause_start_row, column=clause_col_idx + 1)
                    merged_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=0)


        # Adjust column widths
        column_widths = { # Approximate widths, can be fine-tuned
            headers[0]: 15,  # Clause ID
            headers[1]: 50,  # Clause Title
            headers[2]: 20,  # Requires Procedure
            headers[3]: 25,  # Clause Compliance Status
            headers[4]: 40,  # Clause Compliance Desc
            headers[5]: 40,  # Clause Impr Sugg
            headers[6]: 15,  # Task ID
            headers[7]: 50,  # Task Sentence
            headers[8]: 40,  # Evidence Source
            headers[9]: 10,  # Evidence Page
            headers[10]: 10, # Evidence Score
            headers[11]: 60  # Evidence Excerpt
        }
        for i, header_name in enumerate(headers):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i + 1)].width = column_widths.get(header_name, 20) # Default width 20

        try:
            # 新增外部法規原文 sheet
            if hasattr(self.project, 'external_regulations_json_path') and self.project.external_regulations_json_path and self.project.external_regulations_json_path.exists():
                ws_ext = wb.create_sheet(self.translator.get("excel_sheet_title_external_regulations", "External Regulations"))
                try:
                    with open(self.project.external_regulations_json_path, 'r', encoding='utf-8') as f:
                        ext_content = f.read()
                    # 格式化 JSON
                    try:
                        ext_json = json.loads(ext_content)
                        ext_content = json.dumps(ext_json, indent=2, ensure_ascii=False)
                    except Exception:
                        pass
                    # 標題
                    ws_ext.cell(row=1, column=1, value=self.translator.get("excel_external_regulations_content_col", "External Regulations JSON Content"))
                    ws_ext.cell(row=1, column=1).font = header_font
                    ws_ext.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    ws_ext.cell(row=1, column=1).border = thin_border
                    # 內容
                    lines = ext_content.splitlines()
                    for i, line in enumerate(lines, start=2):
                        cell = ws_ext.cell(row=i, column=1, value=line)
                        cell.font = cell_font
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.border = thin_border
                    ws_ext.column_dimensions['A'].width = 120
                except Exception as e:
                    ws_ext.cell(row=1, column=1, value=self.translator.get("excel_error_reading_file", "Error reading file: {error}").format(error=e))
                    ws_ext.cell(row=1, column=1).font = cell_font
                    ws_ext.cell(row=1, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                    ws_ext.cell(row=1, column=1).border = thin_border

            # 新增內部程序原文 sheet
            if hasattr(self.project, 'procedure_doc_paths') and self.project.procedure_doc_paths:
                ws_proc = wb.create_sheet(self.translator.get("excel_sheet_title_procedures", "Procedures"))
                ws_proc.cell(row=1, column=1, value=self.translator.get("excel_procedure_filename_col", "File Name"))
                ws_proc.cell(row=1, column=2, value=self.translator.get("excel_procedure_content_col", "Content"))
                ws_proc.cell(row=1, column=1).font = header_font
                ws_proc.cell(row=1, column=2).font = header_font
                ws_proc.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_proc.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws_proc.cell(row=1, column=1).border = thin_border
                ws_proc.cell(row=1, column=2).border = thin_border
                ws_proc.column_dimensions['A'].width = 40
                ws_proc.column_dimensions['B'].width = 120
                for idx, proc_path in enumerate(self.project.procedure_doc_paths, start=2):
                    ws_proc.cell(row=idx, column=1, value=str(proc_path.name))
                    ws_proc.cell(row=idx, column=1).font = cell_font
                    ws_proc.cell(row=idx, column=1).alignment = Alignment(vertical="top", wrap_text=True)
                    ws_proc.cell(row=idx, column=1).border = thin_border
                    try:
                        with open(proc_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                        ws_proc.cell(row=idx, column=2, value=content)
                        ws_proc.cell(row=idx, column=2).font = cell_font
                        ws_proc.cell(row=idx, column=2).alignment = Alignment(vertical="top", wrap_text=True)
                        ws_proc.cell(row=idx, column=2).border = thin_border
                    except Exception as e:
                        ws_proc.cell(row=idx, column=2, value=self.translator.get("excel_error_reading_file", "Error reading file: {error}").format(error=e))
                        ws_proc.cell(row=idx, column=2).font = cell_font
                        ws_proc.cell(row=idx, column=2).alignment = Alignment(vertical="top", wrap_text=True)
                        ws_proc.cell(row=idx, column=2).border = thin_border

            wb.save(str(target_file_path))
            QMessageBox.information(self,
                                    self.translator.get("excel_exported_dialog_title", "Excel Exported"),
                                    self.translator.get("excel_exported_dialog_message", "Results successfully exported to: {filepath}").format(filepath=target_file_path))
        except IOError as e:
            logger.error(f"IOError exporting Excel for project {self.project.name} to {target_file_path}: {e}", exc_info=True)
            QMessageBox.critical(self,
                                 self.translator.get("export_error_dialog_title", "Export Error"),
                                 self.translator.get("export_error_dialog_message_excel", "Could not write Excel file: {error}\n\nPlease try saving to a different location or ensure the file is not open elsewhere.").format(error=e))
        except Exception as e:
            logger.error(f"Unexpected error during Excel export for project {self.project.name}: {e}", exc_info=True)
            QMessageBox.critical(self,
                                 self.translator.get("export_error_dialog_title", "Export Error"),
                                 self.translator.get("unexpected_export_error_dialog_message", "An unexpected error occurred during Excel export: {error}").format(error=e))

